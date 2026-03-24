"""Tests for workout_parser.parser."""

from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook

from workout_parser.parser import (
    EXERCISE_PATTERN,
    Exercise,
    _normalize_values,
    first_monday,
    parse_folder,
    parse_month_year,
    parse_workbook,
    write_output,
)


class TestFirstMonday:
    def test_month_starting_on_monday(self):
        # 2024-01-01 is a Monday
        assert first_monday(2024, 1) == date(2024, 1, 1)

    def test_month_starting_on_wednesday(self):
        # 2024-05-01 is a Wednesday -> first Monday is May 6
        assert first_monday(2024, 5) == date(2024, 5, 6)

    def test_month_starting_on_sunday(self):
        # 2024-09-01 is a Sunday -> first Monday is Sep 2
        assert first_monday(2024, 9) == date(2024, 9, 2)


class TestParseMonthYear:
    def test_standard_filename(self):
        assert parse_month_year("December-2024-4-Day-Full-Gym-Routine.xlsx") == (2024, 12)

    def test_short_filename(self):
        assert parse_month_year("May-2024-3-Day-Program.xlsx") == (2024, 5)

    def test_invalid_filename(self):
        with pytest.raises(ValueError, match="Cannot parse month/year"):
            parse_month_year("bad-filename.xlsx")


class TestExercisePattern:
    def test_sets_reps_weight(self):
        m = EXERCISE_PATTERN.match("3x10x135")
        assert m is not None
        assert m.group(1) == "3"
        assert m.group(2) == "10"
        assert m.group(3) == "135"

    def test_sets_reps_only(self):
        m = EXERCISE_PATTERN.match("3x12")
        assert m is not None
        assert m.group(1) == "3"
        assert m.group(2) == "12"
        assert m.group(3) is None

    def test_no_match(self):
        assert EXERCISE_PATTERN.match("not an exercise") is None


class TestNormalizeValues:
    def test_standard_order(self):
        # {sets}x{reps}x{weight} — already correct
        assert _normalize_values(3, 10, 135, prescribed_sets=3, prescribed_reps=(8, 10)) == (3, 10, 135)

    def test_weight_first(self):
        # {weight}x{reps}x{sets} e.g., 45x8x3 with prescribed "3 x 6-8"
        assert _normalize_values(45, 8, 3, prescribed_sets=3, prescribed_reps=(6, 8)) == (3, 8, 45)

    def test_weight_first_swapped_sets_reps(self):
        # {weight}x{sets}x{reps} e.g., 20x3x10 with prescribed "3 x 8-10"
        assert _normalize_values(20, 3, 10, prescribed_sets=3, prescribed_reps=(8, 10)) == (3, 10, 20)

    def test_reps_below_range(self):
        # Reps can be below the prescribed range (user went lower)
        # e.g., 50x5x3 with prescribed "3 x 6-8" — 5 is below range but <= 8
        assert _normalize_values(50, 5, 3, prescribed_sets=3, prescribed_reps=(6, 8)) == (3, 5, 50)

    def test_two_numbers_with_prescribed(self):
        # bodyweight: 3x10
        assert _normalize_values(3, 10, None, prescribed_sets=3) == (3, 10, 0)

    def test_two_numbers_reversed_with_prescribed(self):
        # bodyweight reversed: 30x3 (plank)
        assert _normalize_values(30, 3, None, prescribed_sets=3) == (3, 30, 0)

    def test_fallback_no_prescribed(self):
        # Without prescribed info, falls back to sort heuristic
        assert _normalize_values(45, 8, 3) == (3, 8, 45)

    def test_fallback_two_numbers_no_prescribed(self):
        assert _normalize_values(30, 3, None) == (3, 30, 0)

    def test_prescribed_sets_only_no_reps(self):
        # Has prescribed sets but no rep range — falls back to smaller=reps
        assert _normalize_values(45, 8, 3, prescribed_sets=3) == (3, 8, 45)

    def test_equal_values(self):
        assert _normalize_values(5, 5, 5, prescribed_sets=5, prescribed_reps=(5, 5)) == (5, 5, 5)


def _create_test_workbook(path: Path) -> None:
    """Create a minimal workout Excel file for testing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Week 1"

    # Row 1: Day header
    ws.cell(row=1, column=2, value="Day 1")
    ws.cell(row=1, column=6, value="")

    # Row 2: Exercise
    ws.cell(row=2, column=2, value="A1: Bench Press")
    ws.cell(row=2, column=6, value="3x10x135")

    # Row 3: Another exercise
    ws.cell(row=3, column=2, value="A2: Incline DB Press")
    ws.cell(row=3, column=6, value="3x12x50")

    # Week 2 sheet
    ws2 = wb.create_sheet("Week 2")
    ws2.cell(row=1, column=2, value="Day 1")
    ws2.cell(row=1, column=6, value="")
    ws2.cell(row=2, column=2, value="B1: Squat")
    ws2.cell(row=2, column=6, value="4x8x225")

    # Non-week sheet (should be skipped)
    ws3 = wb.create_sheet("Notes")
    ws3.cell(row=1, column=1, value="Some notes")

    wb.save(path)


class TestParseWorkbook:
    def test_parses_exercises(self, tmp_path):
        xlsx = tmp_path / "May-2024-3-Day-Program.xlsx"
        _create_test_workbook(xlsx)

        exercises = parse_workbook(xlsx)

        assert len(exercises) == 3
        assert exercises[0] == Exercise(
            date=date(2024, 5, 7),  # first Monday May 6 + Day 1
            order="A1",
            name="Bench Press",
            sets=3,
            reps=10,
            weight=135,
        )
        assert exercises[1].name == "Incline DB Press"
        assert exercises[1].weight == 50

        # Week 2 exercise: first Monday + 1 week + Day 1
        assert exercises[2] == Exercise(
            date=date(2024, 5, 14),
            order="B1",
            name="Squat",
            sets=4,
            reps=8,
            weight=225,
        )

    def test_skips_non_week_sheets(self, tmp_path):
        xlsx = tmp_path / "May-2024-3-Day-Program.xlsx"
        _create_test_workbook(xlsx)

        exercises = parse_workbook(xlsx)
        # Should only have exercises from Week 1 and Week 2, not Notes
        assert all(e.name in ("Bench Press", "Incline DB Press", "Squat") for e in exercises)


class TestParseFolder:
    def test_empty_folder(self, tmp_path):
        df = parse_folder(tmp_path)
        assert len(df) == 0
        assert list(df.columns) == ["Date", "Order", "Name", "Sets", "Reps", "Weight"]

    def test_multiple_files(self, tmp_path):
        _create_test_workbook(tmp_path / "May-2024-3-Day-Program.xlsx")
        _create_test_workbook(tmp_path / "September-2024-3-Day-Program-Full-Gym-Routine-2.xlsx")

        df = parse_folder(tmp_path)
        assert len(df) == 6  # 3 exercises per file
        assert df["Date"].is_monotonic_increasing


class TestWriteOutput:
    def test_writes_excel(self, tmp_path):
        xlsx = tmp_path / "May-2024-3-Day-Program.xlsx"
        _create_test_workbook(xlsx)

        df = parse_folder(tmp_path)
        output = tmp_path / "output.xlsx"
        write_output(df, output)

        assert output.exists()
        # Read it back and verify
        import pandas as pd

        result = pd.read_excel(output)
        assert len(result) == 3
        assert list(result.columns) == ["Date", "Order", "Name", "Sets", "Reps", "Weight"]