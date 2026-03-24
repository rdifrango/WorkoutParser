"""Core parsing logic for workout spreadsheet files."""

import re
import tempfile
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import BinaryIO

import pandas as pd
from openpyxl import load_workbook

EXERCISE_PATTERN = re.compile(r"^(\d+)x(\d+)(?:x(\d+))?$")
PRESCRIBED_PATTERN = re.compile(r"^(\d+)\s*x\s*(\d+)(?:\s*-\s*(\d+))?")

DAY_PATTERN = re.compile(r"Day\s+(\d+)")
WEEK_PATTERN = re.compile(r"Week\s+(\d+)")
FILENAME_PATTERN = re.compile(r"^([A-Za-z]+)-(\d{4})-")


@dataclass
class Exercise:
    date: date
    order: str
    name: str
    sets: int
    reps: int
    weight: int


def first_monday(year: int, month: int) -> date:
    """Return the first Monday of the given month."""
    d = date(year, month, 1)
    days_ahead = (7 - d.weekday()) % 7  # Monday is 0
    return d + timedelta(days=days_ahead)


def parse_month_year(filename: str) -> tuple[int, int]:
    """Extract month and year from a filename like 'December-2024-4-Day-...'."""
    match = FILENAME_PATTERN.match(filename)
    if not match:
        raise ValueError(f"Cannot parse month/year from filename: {filename}")
    month_name, year_str = match.group(1), match.group(2)
    month_num = datetime.strptime(month_name, "%B").month
    return int(year_str), month_num


SUPPORTED_EXTENSIONS = {".xlsx", ".numbers"}


def _normalize_values(
    a: int,
    b: int,
    c: int | None,
    prescribed_sets: int | None = None,
    prescribed_reps: tuple[int, int] | None = None,
) -> tuple[int, int, int]:
    """Normalize parsed numbers into (sets, reps, weight) regardless of input order.

    Uses prescribed_sets and prescribed_reps (low, high) from the Sets/Reps column
    to identify which numbers are sets, reps, and weight.
    Falls back to a sort-based heuristic when prescribed info is not available.
    """
    if c is None:
        # Two-number entry: bodyweight exercise, no weight
        if prescribed_sets is not None and prescribed_sets in (a, b):
            reps = b if a == prescribed_sets else a
            return prescribed_sets, reps, 0
        lo, hi = sorted([a, b])
        return lo, hi, 0

    nums = [a, b, c]

    if prescribed_sets is not None:
        # Remove the first occurrence matching prescribed_sets
        try:
            nums.remove(prescribed_sets)
        except ValueError:
            pass
        else:
            # Use rep range to identify reps vs weight
            if prescribed_reps is not None:
                lo, hi = prescribed_reps
                for i, n in enumerate(nums):
                    if lo <= n <= hi or n <= hi:
                        reps = n
                        weight = nums[1 - i]
                        return prescribed_sets, reps, weight
            # Fallback: smaller = reps, larger = weight
            reps, weight = sorted(nums)
            return prescribed_sets, reps, weight

    # Fallback: smallest = sets, middle = reps, largest = weight
    nums = sorted([a, b, c])
    return nums[0], nums[1], nums[2]


def _parse_rows(rows, week_date: date) -> list[Exercise]:
    """Extract exercises from an iterable of row values for a single week sheet."""
    exercises: list[Exercise] = []
    current_date = week_date

    for row in rows:
        exercise_cell = row[1] if len(row) > 1 else None
        prescribed_cell = row[2] if len(row) > 2 else None
        set_cell = row[5] if len(row) > 5 else None

        if not exercise_cell:
            continue

        exercise_val = str(exercise_cell).strip()

        day_match = DAY_PATTERN.match(exercise_val)
        if day_match:
            current_date = week_date + timedelta(days=int(day_match.group(1)))

        if not set_cell:
            continue

        set_val = str(set_cell).strip()
        set_match = EXERCISE_PATTERN.match(set_val)
        if set_match:
            if ":" not in exercise_val:
                continue
            order, name = exercise_val.split(":", 1)

            # Extract prescribed sets and rep range from column C
            # e.g., "3 x 6-8" → sets=3, reps=(6, 8)
            prescribed_sets = None
            prescribed_reps = None
            if prescribed_cell:
                ps_match = PRESCRIBED_PATTERN.match(str(prescribed_cell).strip())
                if ps_match:
                    prescribed_sets = int(ps_match.group(1))
                    rep_lo = int(ps_match.group(2))
                    rep_hi = int(ps_match.group(3)) if ps_match.group(3) else rep_lo
                    prescribed_reps = (rep_lo, rep_hi)

            raw_c = int(set_match.group(3)) if set_match.group(3) else None
            sets, reps, weight = _normalize_values(
                int(set_match.group(1)), int(set_match.group(2)), raw_c,
                prescribed_sets, prescribed_reps,
            )
            exercises.append(
                Exercise(
                    date=current_date,
                    order=order.strip(),
                    name=name.strip(),
                    sets=sets,
                    reps=reps,
                    weight=weight,
                )
            )

    return exercises


def parse_workbook(path: Path) -> list[Exercise]:
    """Parse a single workout spreadsheet and return a list of exercises."""
    suffix = path.suffix.lower()
    if suffix == ".numbers":
        return _parse_numbers_workbook(path)
    return _parse_excel_workbook(path)


def _parse_excel_workbook(path: Path) -> list[Exercise]:
    """Parse a single workout Excel file."""
    year, month = parse_month_year(path.name)
    monday = first_monday(year, month)
    exercises: list[Exercise] = []

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            week_match = WEEK_PATTERN.match(sheet_name)
            if not week_match:
                continue

            week_num = int(week_match.group(1))
            week_date = monday + timedelta(weeks=week_num - 1)

            ws = wb[sheet_name]
            rows = ([cell.value for cell in row] for row in ws.iter_rows(min_row=1))
            exercises.extend(_parse_rows(rows, week_date))
    finally:
        wb.close()

    return exercises


def _parse_numbers_workbook(path: Path) -> list[Exercise]:
    """Parse a single workout Numbers file."""
    from numbers_parser import Document

    year, month = parse_month_year(path.name)
    monday = first_monday(year, month)
    exercises: list[Exercise] = []

    doc = Document(str(path))
    for sheet in doc.sheets:
        week_match = WEEK_PATTERN.match(sheet.name)
        if not week_match:
            continue

        week_num = int(week_match.group(1))
        week_date = monday + timedelta(weeks=week_num - 1)

        table = sheet.tables[0]
        rows = (
            [table.cell(row, col).value for col in range(table.num_cols)]
            for row in range(table.num_rows)
        )
        exercises.extend(_parse_rows(rows, week_date))

    return exercises


def validate_file(filename: str) -> str | None:
    """Validate that a filename matches the expected workout format.

    Returns an error message string, or None if valid.
    """
    match = FILENAME_PATTERN.match(filename)
    if not match:
        return (
            f"'{filename}' doesn't match expected format "
            "(e.g., 'December-2024-4-Day-Full-Gym-Routine.xlsx' or .numbers)"
        )
    month_name = match.group(1)
    try:
        datetime.strptime(month_name, "%B")
    except ValueError:
        return f"'{filename}' has an invalid month name: '{month_name}'"
    return None


def parse_folder(folder: Path) -> pd.DataFrame:
    """Parse all workout spreadsheet files in a folder and return a consolidated DataFrame."""
    all_exercises: list[Exercise] = []
    for path in sorted(
        p for ext in SUPPORTED_EXTENSIONS for p in folder.glob(f"*{ext}")
    ):
        all_exercises.extend(parse_workbook(path))

    if not all_exercises:
        return pd.DataFrame(columns=["Date", "Order", "Name", "Sets", "Reps", "Weight"])

    df = pd.DataFrame(
        [
            {
                "Date": e.date,
                "Order": e.order,
                "Name": e.name,
                "Sets": e.sets,
                "Reps": e.reps,
                "Weight": e.weight,
            }
            for e in all_exercises
        ]
    )
    df = normalize_names(df)
    return df.sort_values("Date").reset_index(drop=True)


def parse_files(files: list[BinaryIO]) -> pd.DataFrame:
    """Parse uploaded file objects and return a consolidated DataFrame."""
    all_exercises: list[Exercise] = []
    for f in files:
        original_name = getattr(f, "name", "upload.xlsx")
        suffix = Path(original_name).suffix or ".xlsx"
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
            tmp.write(f.read())
            tmp_path = Path(tmp.name)
        # Use the original filename for month/year parsing
        tmp_renamed = tmp_path.parent / original_name
        tmp_path.rename(tmp_renamed)
        try:
            all_exercises.extend(parse_workbook(tmp_renamed))
        finally:
            tmp_renamed.unlink(missing_ok=True)

    if not all_exercises:
        return pd.DataFrame(columns=["Date", "Order", "Name", "Sets", "Reps", "Weight"])

    df = pd.DataFrame(
        [
            {
                "Date": e.date,
                "Order": e.order,
                "Name": e.name,
                "Sets": e.sets,
                "Reps": e.reps,
                "Weight": e.weight,
            }
            for e in all_exercises
        ]
    )
    df = normalize_names(df)
    return df.sort_values("Date").reset_index(drop=True)


PAREN_SUFFIX = re.compile(r"\s*\([^)]*\)\s*$")
ANGLE_OR_SUFFIX = re.compile(r"\s*<\s*or\s*>.*$")


def normalize_names(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize exercise names to reduce duplicates from inconsistent naming."""
    if df.empty:
        return df

    # Strip trailing parenthetical aliases, e.g., "Dumbbell Romanian Deadlift (RDL)"
    df["Name"] = df["Name"].str.replace(PAREN_SUFFIX, "", regex=True)
    # Strip "< or > ..." alternatives, e.g., "Barbell Bench Press < or > Machine Chest Press"
    df["Name"] = df["Name"].str.replace(ANGLE_OR_SUFFIX, "", regex=True)

    # Build a mapping for names that differ only by a trailing 's'.
    # Keep the more frequent form as canonical.
    counts = df["Name"].value_counts()
    name_map: dict[str, str] = {}
    seen: set[str] = set()

    for name in counts.index:
        if name in seen:
            continue
        # Check if singular/plural counterpart exists
        if name.endswith("s"):
            other = name[:-1]
        else:
            other = name + "s"

        if other in counts.index and other not in seen:
            # Keep whichever has more occurrences
            canonical = name if counts[name] >= counts[other] else other
            variant = other if canonical == name else name
            name_map[variant] = canonical
            seen.update([name, other])

    if name_map:
        df["Name"] = df["Name"].replace(name_map)

    return df


def write_output(df: pd.DataFrame, output_path: Path) -> None:
    """Write the consolidated DataFrame to an Excel file."""
    df.to_excel(output_path, index=False, sheet_name="Monthly Exercises")